import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from datetime import datetime

OUT = r"C:\Users\edgar.suarez\Downloads\Control Vendors ATIP CSE.xlsx"

# === Source rows from "Asistencias Tecnicas CSE 18ABR2026.xlsx" rows 9-37 ===
# (admin, contrato_marco, ods, acta_inicio, vigencia, duracion, subproyecto, vendor, alcance, DT, DE, DR, acum, vr_unit_dia)
def d(s):
    if s is None or s == "TBD" or s == "":
        return None
    return s

V = []
def add(*args):
    V.append(args)

add("Sebastian Arias","3032969-2021-005","TBD","TBD","2026-06-30",372,"CLUSTER EXTERIORES",
    "SCHNEIDER ELECTRIC DE COLOMBIA S.A",
    "Asistencia técnica para el comisionamiento, arranque y estabilización de equipos",
    372,361,11,197,None)
add("Sebastian Arias","3032928-2021-009","TBD","TBD","2026-06-30",647,"CLUSTER EXTERIORES",
    "SLA COL S.A.S",
    "Asistencia técnica para el comisionamiento, arranque y estabilización de equipos",
    669,615,54,370,2954647)
add("Wilfredo Charry","CW233230","9996257","TBD","2026-06-30",234,"CLUSTER EXTERIORES",
    "SLA COL S.A.S",
    "Asistencia técnica, pruebas en sitio y capacitación de software",
    234,69,165,None,None)
add("Andres Pereira Castaño","5208529","TBD","TBD","12 días",2,"MITO 2 ESTACIONES",
    "TRIENERGY",
    "Cambio de control y puesta en marcha de sincronismo en generador",
    2,0,2,None,None)
add("Angela Maria Correa","CW191332","TBD","2025-04-02","2026-07-31",982,"CENTAUROS ESTACIONES",
    "ROCKWELL AUTOMATION COLOMBIA S.A.",
    "Sistema de control EP1",
    982,282.5,699.5,188,None)
add("Angela Maria Correa","CW191332","TBD","2025-04-02","2026-07-31",132,"MITO 2 ESTACIONES",
    "ROCKWELL AUTOMATION COLOMBIA S.A.",
    "Sistema de control MITO 2",
    132,75.5,56.5,37,None)
add("Andres Pereira Castaño","5208529","TBD","TBD","TBD",12,"CENTAUROS ESTACIONES",
    "TRIENERGY",
    "Generador trifásico 500 KW a 480/277 V con motor de combustión interna (diésel)",
    2,0,2,None,None)
add("William Andrés Romero","CW191659","9696894","TBD","2025-12-31",12,"CENTAUROS ESTACIONES",
    "FLOWSERVE",
    "Bombas FWKO: retrofit bomba de crudo EP-1150 J + VFD + VMS",
    12,8,4,None,None)
add("Wilfredo Charry","3050816","9758665","2025-11-18","2026-10-22",20,"MITO 2 ESTACIONES",
    "SCHNEIDER ELECTRIC DE COLOMBIA S.A",
    "ODB 8001227: tablero DCS SCADA PAD - suministro de tablero con servicio de asistencia",
    20,12,8,None,None)
add("Wilfredo Charry","3050816","9765021","2025-11-18","2026-10-17",14,"CENTAUROS ESTACIONES",
    "SCHNEIDER ELECTRIC DE COLOMBIA S.A",
    "ODB 8001228: SCADA Eléctrico Centauros",
    14,7,7,3,None)
add("Samuel Orlando Blanco Ruiz","3050619","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "EQUIPOS Y LABORATORIO DE COLOMBIA",
    "MOD3: asistencia técnica de equipos de laboratorio EP1 (CSE)",
    0,0,0,None,None)
add("Samuel Orlando Blanco Ruiz","3050624","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "HACH COLOMBIA S.A.S.",
    "MOD3: asistencia técnica de equipos de laboratorio EP1 (CSE)",
    0,0,0,None,None)
add("Omar Enrique Orjuela","5209050","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "ALSA SOLUCIONES",
    "Servicio de asistencia técnica en campo, proyecto CSE Módulo 3",
    0,0,0,None,None)
add("Samuel Orlando Blanco Ruiz","CW223232","TBD","TBD","TBD",3,"CENTAUROS ESTACIONES",
    "INSURCOL S.A.S",
    "Suministro de detectores y estaciones manuales Fire & Gas para Estación Centauros",
    3,0,3,None,None)
add("Omar Enrique Orjuela","3043959","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "GIM INGENIERIA ELECTRICA LTDA",
    "Asistencia técnica para tableros eléctricos de EP1 (CSE Mód.)",
    0,0,0,None,None)
add("Omar Enrique Orjuela","3047457","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "EQUIPOS Y LABORATORIO DE COLOMBIA",
    "MOD3: asistencia técnica de equipos de laboratorio EP1",
    0,0,0,None,None)
add("Andres Pereira Castaño","3050626","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "S Y Z COLOMBIA",
    "Asistencia técnica de equipos de laboratorio EP1 (CSE Mód.)",
    0,0,0,None,None)
add("Andres Pereira Castaño","3050528","9985674","TBD","2026-10-28",13,"CENTAUROS ESTACIONES",
    "INDUSTRIAS ECTRICOL S.A.S.",
    "Servicio de asistencia CMO-06 CSE M4 de ECOPETROL",
    13,0,13,None,None)
add("Omar Enrique Orjuela","CW220864","5350641","TBD","TBD",12,"CENTAUROS ESTACIONES",
    "EQUIPOS Y CONTROLES INDUSTRIALES",
    "Servicio de asistencia técnica en campo, proyecto Desarrollo Integral CSE",
    12,0,12,None,None)
add("Sebastian Arias","3032970-2025-002","TBD","TBD","2026-03-30",14,"CLUSTER EXTERIORES",
    "HITACHI ENERGY COLOMBIA LTDA",
    "Asistencia técnica para el comisionamiento, arranque y estabilización de equipos",
    14,12,2,None,None)
add("Sebastian Arias","3032851-2025-004","TBD","TBD","2026-06-30",8,"CLUSTER EXTERIORES",
    "WEG COLOMBIA S.A.S",
    "Suministro y asistencia técnica de equipos eléctricos de superficie",
    8,2,6,None,None)
add("Gustavo Adolfo Rhenals","3043145","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "SMART INSTRUMENTS SAS","(sin alcance asignado)",
    0,0,0,None,None)
add("Gustavo Adolfo Rhenals","3047142","TBD","TBD","TBD",None,"CENTAUROS ESTACIONES",
    "PUFFER COLOMBIA","(sin alcance asignado)",
    0,0,0,None,None)
add("Sebastian Arias","CW300916","TBD","2026-02-06","2026-07-15",50,"CENTAUROS ESTACIONES",
    "SCHLUMBERGER SURENCO S.A",
    "Asistencia técnica para la reparación y mejora tecnológica de tratador CPF2",
    50,48.5,1.5,None,None)
add("TBD","CW304616","TBD","TBD","TBD",120,"CENTAUROS ESTACIONES",
    "ROCKWELL AUTOMATION COLOMBIA S.A.",
    "Suministro, instalación, cableado, conexionado, configuración y puesta en marcha",
    22,0,22,None,None)
add("Omaira Lucia Tobar Portilla","CW358905","TBD","TBD","TBD",120,"CENTAUROS ESTACIONES",
    "ROCKWELL AUTOMATION COLOMBIA S.A.",
    "Suministro, configuración, integración, pruebas y puesta en operación de equipos nuevos",
    24,15,9,None,None)
add("Marcela Benavides Bohórquez","CW339416","0","TBD","TBD",7,"CENTAUROS ESTACIONES",
    "EXPERTOS EN SOLUCIONES ELECTRICAS S.A.S",
    "Asistencia técnica para configuración y pruebas de UPS en Estación Centauros",
    7,6,1,None,None)
add("Marcela Benavides Bohórquez","3050816","90441450","2026-02-13","2026-05-16",33,"CENTAUROS ESTACIONES",
    "SCHNEIDER ELECTRIC DE COLOMBIA S.A",
    "ID15839: asistencia técnica para integración celda +s2h06 al sistema de control eléctrico",
    33,0,33,None,None)
add("Diana Rubiela Amaya Ramirez","CW345538","5353346","2026-05-07","2026-11-03",24,"CENTAUROS ESTACIONES",
    "ITT GOULDS",
    "Servicio para la modificación bomba EP1105-7 (cabeza de crudo) y suministro",
    24,0,24,None,None)

# === Build workbook ===
wb = Workbook()
ws = wb.active
ws.title = "Control Vendors"

# Styles
NAVY = "1F3864"; ORANGE = "F39200"; LIGHT = "D6E4F0"; GREEN = "548235"
GRAY = "2F4F4F"; ALERT = "C00000"; BG_HEAD = "1F3864"
WHITE = "FFFFFF"
font_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
font_sub   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_h     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_body  = Font(name="Calibri", size=10, color="000000")
font_kpi_l = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
font_kpi_v = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
font_memo  = Font(name="Calibri", size=10, color="1F3864")
fill_title = PatternFill("solid", fgColor=NAVY)
fill_sub   = PatternFill("solid", fgColor=ORANGE)
fill_head  = PatternFill("solid", fgColor=GRAY)
fill_light = PatternFill("solid", fgColor=LIGHT)
fill_kpi1  = PatternFill("solid", fgColor=NAVY)
fill_kpi2  = PatternFill("solid", fgColor=ORANGE)
fill_kpi3  = PatternFill("solid", fgColor=GREEN)
fill_memo  = PatternFill("solid", fgColor="FFF2CC")
align_c    = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_l    = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_r    = Alignment(horizontal="right", vertical="center", wrap_text=True)
align_lt   = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="808080")
medium = Side(border_style="medium", color=NAVY)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Column widths
widths = {"A":4,"B":22,"C":18,"D":12,"E":12,"F":12,"G":9,"H":11,"I":18,"J":32,"K":36,
          "L":7,"M":7,"N":7,"O":7,"P":8,"Q":16,"R":18,"S":18,"T":18,"U":12,"V":22}
for c,w in widths.items():
    ws.column_dimensions[c].width = w

# Row 1: TITLE
ws.merge_cells("B1:V1")
c = ws["B1"]; c.value = "CONTROL DE VENDORS - ASISTENCIAS TÉCNICAS CSE / ECOPETROL"
c.font = font_title; c.fill = fill_title; c.alignment = align_c
ws.row_dimensions[1].height = 32

# Row 2: subtitle
ws.merge_cells("B2:V2")
c = ws["B2"]; c.value = "Caño Sur Este  ·  Estación Centauros  ·  APPLUS+ Colombia  ·  Última actualización: " + datetime.today().strftime("%Y-%m-%d")
c.font = font_sub; c.fill = fill_sub; c.alignment = align_c
ws.row_dimensions[2].height = 20

# Row 4-5: KPI cards
ws.row_dimensions[4].height = 16
ws.row_dimensions[5].height = 28
kpis = [
    ("VENDORS TOTALES", f"=COUNTA(B9:B{8+len(V)})", fill_kpi1),
    ("DÍAS CONTRATADOS", f"=SUM(L9:L{8+len(V)})", fill_kpi1),
    ("DÍAS EJECUTADOS", f"=SUM(M9:M{8+len(V)})", fill_kpi3),
    ("DÍAS PENDIENTES", f"=SUM(N9:N{8+len(V)})", fill_kpi2),
    ("% EJECUCIÓN GLOBAL", f"=IFERROR(SUM(M9:M{8+len(V)})/SUM(L9:L{8+len(V)}),0)", fill_kpi3),
    ("VALOR PENDIENTE (COP)", f"=SUMPRODUCT(N9:N{8+len(V)},Q9:Q{8+len(V)})", fill_kpi2),
]
col_start = 2  # B
span = 3
for i,(label,formula,fill) in enumerate(kpis):
    c0 = col_start + i*span
    L0 = get_column_letter(c0); L1 = get_column_letter(c0+span-1)
    ws.merge_cells(f"{L0}4:{L1}4")
    ws.merge_cells(f"{L0}5:{L1}5")
    lab = ws[f"{L0}4"]; lab.value = label; lab.font = font_kpi_l; lab.fill = fill; lab.alignment = align_c
    val = ws[f"{L0}5"]; val.value = formula; val.font = font_kpi_v; val.fill = fill; val.alignment = align_c
    if "%" in label:
        val.number_format = "0.0%"
    elif "COP" in label or "VALOR" in label:
        val.number_format = '"$"#,##0'
    else:
        val.number_format = "#,##0"
ws.row_dimensions[4].height = 18
ws.row_dimensions[5].height = 32

# Row 7-8: Headers (2-row header)
ws.row_dimensions[7].height = 8
HEADERS = [
    ("B","ADMINISTRADOR"),("C","VENDOR"),("D","CONTRATO MARCO"),("E","ODS"),
    ("F","ACTA INICIO"),("G","VIGENCIA"),("H","SUBPROYECTO"),
    ("I","TIPO ÁREA"),("J","EQUIPO / SISTEMA"),("K","ALCANCE"),
    ("L","DT\nDías Totales"),("M","DE\nDías Ejec."),("N","DR\nDías Restantes"),
    ("O","ACUM"),("P","% EJEC"),
    ("Q","Vr UNIT/DÍA (COP)"),("R","Vr TOTAL CONTRATO"),("S","Vr EJECUTADO"),
    ("T","Vr PENDIENTE"),("U","ESTADO"),("V","OBSERVACIONES")
]
for col,txt in HEADERS:
    c = ws[f"{col}8"]; c.value = txt
    c.font = font_h; c.fill = fill_head; c.alignment = align_c; c.border = border
ws.row_dimensions[8].height = 38

# Body rows
start = 9
for i, row in enumerate(V):
    r = start + i
    (admin, cm, ods, acta, vig, dur, sub, vendor, alcance, DT, DE, DR, acum, vrunit) = row
    cells = {
        "B": admin,
        "C": vendor,
        "D": cm,
        "E": ods,
        "F": acta,
        "G": vig,
        "H": "EXTERIORES" if "EXT" in sub else ("MITO 2" if "MITO" in sub else "CENTAUROS"),
        "I": sub.replace("CLUSTER EXTERIORES","CLUSTERS").replace("CENTAUROS ESTACIONES","ESTACIÓN CENTAUROS").replace("MITO 2 ESTACIONES","MITO 2"),
        "J": "",  # equipo placeholder
        "K": alcance,
        "L": DT,
        "M": DE,
        "N": DR,
        "O": acum,
        # P = formula
        "Q": vrunit,
        # R, S, T = formulas
        "U": None,  # state formula below
        "V": None,
    }
    for col, v in cells.items():
        c = ws[f"{col}{r}"]
        c.value = v
        c.font = font_body
        c.alignment = align_lt if col in ("K","J","V") else align_c
        c.border = border
    # Formulas
    ws[f"P{r}"] = f"=IFERROR(M{r}/L{r},0)"; ws[f"P{r}"].number_format = "0.0%"
    ws[f"R{r}"] = f"=IFERROR(L{r}*Q{r},\"\")"; ws[f"R{r}"].number_format = '"$"#,##0'
    ws[f"S{r}"] = f"=IFERROR(M{r}*Q{r},\"\")"; ws[f"S{r}"].number_format = '"$"#,##0'
    ws[f"T{r}"] = f"=IFERROR(N{r}*Q{r},\"\")"; ws[f"T{r}"].number_format = '"$"#,##0'
    ws[f"Q{r}"].number_format = '"$"#,##0'
    # Estado
    ws[f"U{r}"] = (
        f'=IF(L{r}=0,"POR INICIAR",IF(M{r}=L{r},"FINALIZADO",IF(M{r}=0,"POR INICIAR","EN EJECUCIÓN")))'
    )
    # alternate row shading
    if i % 2 == 1:
        for col,_ in HEADERS:
            ws[f"{col}{r}"].fill = fill_light
    # auto-fit row height by text length in K
    txtlen = len(str(alcance) or "")
    h = max(20, min(60, 18 + (txtlen // 60) * 12))
    ws.row_dimensions[r].height = h

# Totals row
tot_row = start + len(V)
ws[f"B{tot_row}"] = "TOTALES"
ws.merge_cells(f"B{tot_row}:K{tot_row}")
c = ws[f"B{tot_row}"]; c.font = Font(bold=True, size=11, color="FFFFFF")
c.fill = fill_head; c.alignment = align_r; c.border = border
for col in ["B","C","D","E","F","G","H","I","J","K"]:
    ws[f"{col}{tot_row}"].fill = fill_head
    ws[f"{col}{tot_row}"].border = border
for col, formula, fmt in [
    ("L", f"=SUM(L9:L{tot_row-1})", "#,##0"),
    ("M", f"=SUM(M9:M{tot_row-1})", "#,##0"),
    ("N", f"=SUM(N9:N{tot_row-1})", "#,##0"),
    ("O", f"=SUM(O9:O{tot_row-1})", "#,##0"),
    ("P", f"=IFERROR(M{tot_row}/L{tot_row},0)", "0.0%"),
    ("R", f"=SUM(R9:R{tot_row-1})", '"$"#,##0'),
    ("S", f"=SUM(S9:S{tot_row-1})", '"$"#,##0'),
    ("T", f"=SUM(T9:T{tot_row-1})", '"$"#,##0'),
]:
    cc = ws[f"{col}{tot_row}"]; cc.value = formula; cc.number_format = fmt
    cc.font = Font(bold=True, size=11, color="FFFFFF"); cc.fill = fill_head
    cc.alignment = align_c; cc.border = border
ws.row_dimensions[tot_row].height = 24

# Conditional formatting on % EJEC and Estado
n_end = tot_row - 1
ws.conditional_formatting.add(
    f"P9:P{n_end}",
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B'))
ws.conditional_formatting.add(
    f"U9:U{n_end}",
    FormulaRule(formula=[f'$U9="FINALIZADO"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))
ws.conditional_formatting.add(
    f"U9:U{n_end}",
    FormulaRule(formula=[f'$U9="EN EJECUCIÓN"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C5700", bold=True)))
ws.conditional_formatting.add(
    f"U9:U{n_end}",
    FormulaRule(formula=[f'$U9="POR INICIAR"'], fill=PatternFill("solid", fgColor="F8CBAD"), font=Font(color="833C0C", bold=True)))

# Freeze panes (after headers)
ws.freeze_panes = "C9"
# Autofilter
ws.auto_filter.ref = f"B8:V{tot_row-1}"

# === Memo block ===
memo_row = tot_row + 3
ws.merge_cells(f"B{memo_row}:V{memo_row}")
c = ws[f"B{memo_row}"]; c.value = "📋  BLOQUE PARA COPIAR/PEGAR EN EL MEMORANDO"
c.font = Font(bold=True, size=14, color="FFFFFF"); c.fill = fill_title; c.alignment = align_c
ws.row_dimensions[memo_row].height = 26

memo_lines = [
    ("RESUMEN GENERAL DEL CONTRATO",""),
    ("Vendors activos:",            f"=COUNTIF(U9:U{n_end},\"EN EJECUCIÓN\")"),
    ("Vendors por iniciar:",        f"=COUNTIF(U9:U{n_end},\"POR INICIAR\")"),
    ("Vendors finalizados:",        f"=COUNTIF(U9:U{n_end},\"FINALIZADO\")"),
    ("Total días contratados:",     f"=SUM(L9:L{n_end})"),
    ("Total días ejecutados:",      f"=SUM(M9:M{n_end})"),
    ("Total días pendientes:",      f"=SUM(N9:N{n_end})"),
    ("% Ejecución global:",         f"=IFERROR(SUM(M9:M{n_end})/SUM(L9:L{n_end}),0)"),
    ("Valor total contratado:",     f"=SUM(R9:R{n_end})"),
    ("Valor ejecutado:",            f"=SUM(S9:S{n_end})"),
    ("Valor pendiente por facturar:",f"=SUM(T9:T{n_end})"),
    ("",""),
    ("SLA COL S.A.S - ODS 3032928-2021-009  (Memorando #17)",""),
    ("Días totales SLA COL:",       f"=SUMIF(C9:C{n_end},\"SLA COL*\",L9:L{n_end})"),
    ("Días ejecutados SLA COL:",    f"=SUMIF(C9:C{n_end},\"SLA COL*\",M9:M{n_end})"),
    ("Días pendientes SLA COL:",    f"=SUMIF(C9:C{n_end},\"SLA COL*\",N9:N{n_end})"),
    ("Valor pendiente SLA COL:",    f"=SUMPRODUCT((C9:C{n_end}=\"SLA COL S.A.S\")*N9:N{n_end}*Q9:Q{n_end})"),
]
for i, (label, value) in enumerate(memo_lines):
    r = memo_row + 1 + i
    ws[f"B{r}"] = label
    ws[f"B{r}"].font = Font(bold=label.endswith(":") is False and bool(label), size=11, color="1F3864") if label and not label.endswith(":") else Font(bold=False, size=10, color="000000")
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
    if value != "":
        ws[f"I{r}"] = value
        ws.merge_cells(f"I{r}:K{r}")
        # format
        if "%" in label:
            ws[f"I{r}"].number_format = "0.0%"
        elif "Valor" in label or "$" in label:
            ws[f"I{r}"].number_format = '"$"#,##0'
        else:
            ws[f"I{r}"].number_format = "#,##0"
        ws[f"I{r}"].font = Font(bold=True, size=11, color="C00000")
        ws[f"I{r}"].alignment = Alignment(horizontal="right", vertical="center")
    # section header rows
    if label and not value and not label.endswith(":"):
        ws[f"B{r}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"B{r}"].fill = fill_sub
        ws[f"B{r}"].alignment = align_c
        ws.merge_cells(f"B{r}:K{r}")
    ws.row_dimensions[r].height = 18

# Save
wb.save(OUT)
print("OK ->", OUT)
print("Size:", os.path.getsize(OUT))
